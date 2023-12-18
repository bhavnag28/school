from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, views
from .serializers import StudentCreateSerializer
from django.template.loader import render_to_string
from django.core.mail import EmailMessage
from django.conf import settings
from .models import *
from .serializers import *
import pandas as pd
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.pagination import PageNumberPagination
from reportlab.pdfgen import canvas
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

""" Curd API of  student, Parent, Academic details and documents """
class StudentCreateAPIView(APIView):
    pagination_class = PageNumberPagination
    page_size = 10  # Set the number of items you want to display per page   
    
    def send_student_enrollment_email(self, student):
        name = student.name
        enrollment_id = student.academicdetails.enrollment_id  # Adjust the related name here

        email = student.mail_id

        email_subject = f"Enrollment id : {name}"
        context = {
            "email": email,
            "name": name,
            "enrollment_id": enrollment_id
        }
        email_message = render_to_string('mail/studentemail.html', context)
        to_email = [email]

        email = EmailMessage(
            subject=email_subject,
            body=email_message,
            to=to_email,
        )
        email.send()

    def post(self, request, *args, **kwargs):
        serializer = StudentCreateSerializer(data=request.data)
        if serializer.is_valid():
            student = serializer.save()

            # Send email to the student
            self.send_student_enrollment_email(student)

            # Send notification email to a specific email address
            notification_subject = "New Student Enrollment Notification"
            notification_body = f"A new student has been enrolled with id {student.id}."
            notification_email = EmailMessage(
                subject=notification_subject,
                body=notification_body,
                to=["durganand.jha@habrie.com"],
            )
            notification_email.send()

            response_data = {
                "status": "success",
                "message": "Student created successfully",
                "data": {
                    "student_id": student.id
                }
            }

            return Response(response_data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
     
    def get(self, request, *args, **kwargs):
        academicdetails = AcademicDetails.objects.all()

        clas_name = self.request.GET.get('class_name', None)
        section_name = self.request.GET.get('section', None)
        if clas_name and section_name:
            academicdetails = academicdetails.filter(class_name=clas_name, section=section_name)

        students = Student.objects.filter(id__in=academicdetails.values_list('student',flat=True))
        # Apply pagination

        paginator = self.pagination_class()
        result_page = paginator.paginate_queryset(students, request)
        serializer = StudentSerializer(result_page, many=True)
        return paginator.get_paginated_response(serializer.data)

    def put(self, request, pk, format=None):
        student = get_object_or_404(Student, pk=pk)
        serializer = StudentCreateSerializer(student, data=request.data)
        if serializer.is_valid():
            serializer.save()

            # Send email to the student
            self.send_student_enrollment_email(student)
            # Send notification email to a specific email address
            notification_subject = "New Student Enrollment Notification"
            notification_body = f"A new student has been enrolled with id {student.id}."
            notification_email = EmailMessage(
                subject=notification_subject,
                body=notification_body,
                to=["durganand.jha@habrie.com"],
            )
            notification_email.send()
            response_data = {
                "status": "success",
                "message": "Student updated successfully",
                "data": {
                    "student_id": student.id
                }
            }

            return Response(response_data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk, format=None):
        student = get_object_or_404(Student, pk=pk)
        student.delete()

        response_data = {
            "status": "success",
            "message": "Student deleted successfully",
            "data": {
                "student_id": pk
            }
        }

        return Response(response_data, status=status.HTTP_204_NO_CONTENT)

 
    

    
""" Excel import """    
class ExcelImportAPIView(views.APIView):
    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')

        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        # Check if the file type is allowed (Excel)
        allowed_extensions = ('.xlsx', '.xls')
        if not file.name.lower().endswith(allowed_extensions):
            return Response({'error': 'Invalid file format. Only Excel (.xlsx, .xls) allowed.'},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            # Read the Excel file into a DataFrame
            df = pd.read_excel(file)

            # Process the DataFrame and create/update records in the database
            for index, row in df.iterrows():
                student_data = {
                    'name': row['name'],
                    'gender': row['gender'],
                    'adhar_card_number': row['adhar_card_no'],
                    'dob': row['dob'],
                    'identification_marks': row['identification_marks'],
                    'height': row['height'],
                    'weight': row['weight'],
                    'mail_id': row['mail_id'],
                    'contact_detail': row['contact_detail'],
                    'address': row['address'],
                }

                # Create or update Student instance
                student, created = Student.objects.update_or_create(mail_id=row['Email'], defaults=student_data)

                # Create Parent instance
                parent_data = {
                    'student': student,
                    'father_name': row['father_name'],
                    'father_qualification': row['father_qualification'],
                    'father_profession': row['father_profession'],
                    'father_designation': row['father_designation'],
                    'father_aadhar_card': row['father_aadhar_card'],
                    'father_mobile_number': row['father_mobile_number'],
                    'father_mail_id': row['father_mail_id'],
                    'mother_name': row['mother_name'],
                    'mother_qualification': row['mother_qualification'],
                    'mother_profession': row['mother_profession'],
                    'mother_designation': row['mother_designation'],
                    'mother_aadhar_card': row['mother_aadhar_card'],
                    'mother_mobile_number': row['mother_mobile_number'],
                    'mother_mail_id': row['mother_mail_id'],
                }
                parent, parent_created = Parent.objects.update_or_create(student=student, defaults=parent_data)

                # Create AcademicDetails instance
                academic_details_data = {
                    'student': student,
                    'enrollment_id': row['enrollment_id'],
                    'class_name': row['class_name'],
                    'section': row['section'],
                    'doj': row['doj'],
                }
                academic_details, academic_details_created = AcademicDetails.objects.update_or_create(student=student, defaults=academic_details_data)

                # Create or update Document instance
                document_data = {
                    'student': student,
                    'file': row['document_file'],  # Assuming you have a column named 'document_file' in your Excel file
                }
                document, document_created = Document.objects.update_or_create(student=student, defaults=document_data)

            return Response({'success': 'Data imported successfully'}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
""" Export filtered data as Pdf """        
class StudentPdfView(APIView):
    def get(self, request, *args, **kwargs):
        academicdetails = AcademicDetails.objects.all()

        clas_name = self.request.GET.get('class_name', None)
        section_name = self.request.GET.get('section', None)
        if clas_name and section_name:
            academicdetails = academicdetails.filter(class_name=clas_name, section=section_name)

        students = Student.objects.filter(id__in=academicdetails.values_list('student', flat=True))
        serializer = StudentSerializer(students, many=True)

        # Render HTML template
        html_string = render_to_string('pdf_template.html', {'data': serializer.data})

        # Generate PDF using ReportLab
        pdf_file = self.generate_pdf(html_string)

        # Prepare HTTP response with PDF content
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=students_data.pdf'
        response.write(pdf_file.getvalue())

        return response

    def generate_pdf(self, html_content):
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer)

        # Set up the PDF document
        pdf.drawString(100, 800, "Student Data PDF")

        # Convert HTML content to PDF
        pdf.drawString(100, 750, html_content)

        pdf.showPage()
        pdf.save()

        buffer.seek(0)
        return buffer
    
""" Export filtered data as excel """    
class ExportExcelView(APIView):
    def get(self, request, *args, **kwargs):
        # Get filter parameters from the request
        class_name = self.request.GET.get('class_name', None)
        section_name = self.request.GET.get('section', None)

        # Filter AcademicDetails based on the parameters
        academicdetails = AcademicDetails.objects.all()
        if class_name and section_name:
            academicdetails = academicdetails.filter(class_name=class_name, section=section_name)

        # Filter Students based on AcademicDetails
        students = Student.objects.filter(id__in=academicdetails.values_list('student', flat=True))
        serializer = StudentSerializer(students, many=True)

        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        # Write headers to the worksheet
        headers = [
            'Name', 'Gender', 'Aadhar Card Number', 'Date of Birth', 'Identification Marks',
            'Admission Category', 'Height', 'Weight', 'Email', 'Contact Detail', 'Address'
        ]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header

        # Write data to the worksheet
        for row_num, item in enumerate(serializer.data, 2):
            for col_num, key in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}{row_num}"] = item.get(key, '')

        # Create a response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students_data.xlsx'
        wb.save(response)

        return response